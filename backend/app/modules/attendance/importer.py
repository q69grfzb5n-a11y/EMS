"""Pure attendance-export parser: bytes in, rows + issues out, zero DB access.

Reused identically for dry-run preview and commit so the parse the user
previewed is exactly the parse that gets written — no separate code path to
drift out of sync.
"""

import calendar
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Literal

import openpyxl

REQUIRED_SHEET_NAME = "Employees Time Card Summary Rep"
EXPECTED_HEADERS = [
    "Month",
    "Person No.",
    "Name",
    "Department",
    "Worksite",
    "Sponsor",
    "Present",
    "Off Days",
    "Absent",
    "Leave",
    "Public Holiday",
    "Deduct Min",
    "Over Time",
    "Approved",
    "Pending Approval",
    "Submitted",
    "Approved Over Time",
]

Severity = Literal["error", "warning"]


@dataclass(frozen=True)
class ParsedAttendanceRow:
    row_number: int
    staff_no: str
    name_raw: str
    present: int
    off_days: int
    absent: int
    leave: int
    public_holiday: int
    deduct_min: Decimal
    over_time: Decimal
    approved: int
    pending_approval: int
    submitted: int
    approved_over_time: Decimal


@dataclass(frozen=True)
class RowIssue:
    row_number: int
    staff_no: str | None
    severity: Severity
    message: str


@dataclass(frozen=True)
class ParseResult:
    rows: list[ParsedAttendanceRow]
    issues: list[RowIssue]

    @property
    def has_errors(self) -> bool:
        return any(issue.severity == "error" for issue in self.issues)


def normalize_staff_no(value: object) -> str:
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    try:
        return str(int(float(text)))
    except ValueError:
        return text


def _to_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal(0)
    return Decimal(str(value))


def _to_int(value: int | float | str | None) -> int:
    if value is None:
        return 0
    return int(value)


def parse_attendance_file(
    content: bytes, *, declared_year: int, declared_month: int
) -> ParseResult:
    declared_label = f"{declared_month:02d}-{declared_year:04d}"
    days_in_month = calendar.monthrange(declared_year, declared_month)[1]

    try:
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        return ParseResult(
            rows=[], issues=[RowIssue(0, None, "error", f"Could not read file: {exc}")]
        )

    if REQUIRED_SHEET_NAME not in workbook.sheetnames:
        return ParseResult(
            rows=[],
            issues=[
                RowIssue(
                    0,
                    None,
                    "error",
                    f"Expected sheet '{REQUIRED_SHEET_NAME}' not found "
                    f"(sheets in file: {', '.join(workbook.sheetnames)})",
                )
            ],
        )

    sheet = workbook[REQUIRED_SHEET_NAME]
    header = [
        cell.value
        for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(EXPECTED_HEADERS)))
    ]
    if header != EXPECTED_HEADERS:
        return ParseResult(
            rows=[],
            issues=[
                RowIssue(
                    1, None, "error", f"Header drift: expected {EXPECTED_HEADERS!r}, got {header!r}"
                )
            ],
        )

    rows: list[ParsedAttendanceRow] = []
    issues: list[RowIssue] = []
    seen_staff_no: dict[str, int] = {}

    for excel_row, cells in enumerate(
        sheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS)), start=2
    ):
        vals = [c.value for c in cells]
        if all(v is None for v in vals):
            continue

        month_value, person_no = vals[0], vals[1]
        if person_no is None:
            issues.append(RowIssue(excel_row, None, "error", "Missing Person No."))
            continue
        staff_no = normalize_staff_no(person_no)

        if str(month_value).strip() != declared_label:
            issues.append(
                RowIssue(
                    excel_row,
                    staff_no,
                    "error",
                    f"Row Month {month_value!r} does not match declared period {declared_label!r}",
                )
            )
            continue

        if staff_no in seen_staff_no:
            first_row = seen_staff_no[staff_no]
            issues.append(
                RowIssue(
                    excel_row,
                    staff_no,
                    "error",
                    f"Duplicate Person No. {staff_no!r} (first seen on row {first_row})",
                )
            )
            continue
        seen_staff_no[staff_no] = excel_row

        present = _to_int(vals[6])
        off_days = _to_int(vals[7])
        absent = _to_int(vals[8])
        leave = _to_int(vals[9])
        public_holiday = _to_int(vals[10])

        bucket_sum = present + off_days + absent + leave + public_holiday
        if bucket_sum != days_in_month:
            issues.append(
                RowIssue(
                    excel_row,
                    staff_no,
                    "warning",
                    f"Present+Off Days+Absent+Leave+Public Holiday = {bucket_sum}, "
                    f"expected {days_in_month} days in {declared_label}",
                )
            )

        rows.append(
            ParsedAttendanceRow(
                row_number=excel_row,
                staff_no=staff_no,
                name_raw=str(vals[2]) if vals[2] is not None else "",
                present=present,
                off_days=off_days,
                absent=absent,
                leave=leave,
                public_holiday=public_holiday,
                deduct_min=_to_decimal(vals[11]),
                over_time=_to_decimal(vals[12]),
                approved=_to_int(vals[13]),
                pending_approval=_to_int(vals[14]),
                submitted=_to_int(vals[15]),
                approved_over_time=_to_decimal(vals[16]),
            )
        )

    return ParseResult(rows=rows, issues=issues)
