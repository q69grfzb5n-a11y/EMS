"""Finance Excel export — pure: builds an openpyxl Workbook from already-loaded
domain data, zero DB access of its own. A "Summary" sheet (grand totals per
department) plus one sheet per department listing its payouts, both RTL for
the real Arabic names (PLAN §Phase 8)."""

from dataclasses import dataclass
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FONT = Font(bold=True)


@dataclass(frozen=True)
class PayoutRow:
    department_id: int
    department_code: str
    department_name_en: str
    department_name_ar: str
    staff_no: str
    full_name_en: str | None
    full_name_ar: str
    evaluation_pct: Decimal
    final_amount: Decimal


def group_by_department(rows: list[PayoutRow]) -> dict[int, list[PayoutRow]]:
    """Shared aggregation grouping used by both this module's Excel export and
    reports.service's JSON period summary, so the two don't independently
    reimplement the same 'group by department' logic."""
    by_dept: dict[int, list[PayoutRow]] = {}
    for row in rows:
        by_dept.setdefault(row.department_id, []).append(row)
    return by_dept


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def build_finance_workbook(
    *, run_no: int, year: int, month: int, rows: list[PayoutRow]
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    assert summary_ws is not None
    summary_ws.title = "Summary"
    summary_ws.sheet_view.rightToLeft = True
    summary_ws.append([f"Incentive Payout — Run #{run_no} — {month:02d}/{year}"])
    summary_ws.append([])
    _write_header(summary_ws, ["Department (EN)", "القسم", "Employees", "Total (SAR)"])

    by_dept = group_by_department(rows)
    dept_ids_by_code = sorted(by_dept, key=lambda dept_id: by_dept[dept_id][0].department_code)

    grand_total = Decimal(0)
    for dept_id in dept_ids_by_code:
        dept_rows = by_dept[dept_id]
        dept_total = sum((r.final_amount for r in dept_rows), Decimal(0))
        grand_total += dept_total
        summary_ws.append(
            [
                dept_rows[0].department_name_en,
                dept_rows[0].department_name_ar,
                len(dept_rows),
                float(dept_total),
            ]
        )
    summary_ws.append(["Grand Total", "الإجمالي", len(rows), float(grand_total)])
    for cell in summary_ws[summary_ws.max_row]:
        cell.font = _HEADER_FONT

    for dept_id in dept_ids_by_code:
        dept_rows = by_dept[dept_id]
        ws = wb.create_sheet(title=dept_rows[0].department_code[:31])
        ws.sheet_view.rightToLeft = True
        _write_header(ws, ["Staff No", "Name (EN)", "الاسم", "Evaluation %", "Amount (SAR)"])
        for r in sorted(dept_rows, key=lambda x: x.staff_no):
            ws.append(
                [
                    r.staff_no,
                    r.full_name_en or "",
                    r.full_name_ar,
                    float(r.evaluation_pct) * 100,
                    float(r.final_amount),
                ]
            )

    return wb
