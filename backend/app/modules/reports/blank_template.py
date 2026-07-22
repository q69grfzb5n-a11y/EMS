"""Blank evaluation template generator — pure: builds an openpyxl Workbook
listing a template's real team roster with blank, validated score columns,
ready for a Foreman/Group Head to print or fill in (PLAN §Phase 8)."""

from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class RosterMember:
    staff_no: str
    full_name_en: str | None
    full_name_ar: str


@dataclass(frozen=True)
class CriterionSpec:
    name_en: str
    name_ar: str
    max_marks: int
    input_mode: str


def build_blank_evaluation_workbook(
    *, template_name_en: str, criteria: list[CriterionSpec], roster: list[RosterMember]
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = (template_name_en or "Template")[:31]
    ws.sheet_view.rightToLeft = True

    headers = ["Staff No", "Name (EN)", "الاسم"] + [
        f"{c.name_en} ({c.max_marks})" for c in criteria
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = roster or [RosterMember(staff_no="", full_name_en="", full_name_ar="")]
    for member in rows:
        blanks: list[str | int | None] = [None] * len(criteria)
        ws.append([member.staff_no, member.full_name_en or "", member.full_name_ar, *blanks])

    first_data_row = 2
    last_data_row = first_data_row + len(rows) - 1
    for idx, criterion in enumerate(criteria):
        col_letter = get_column_letter(4 + idx)
        if criterion.input_mode == "scale_1_5":
            validation = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
        else:
            validation = DataValidation(
                type="whole", operator="between", formula1=0, formula2=criterion.max_marks
            )
        validation.error = "Invalid value"
        ws.add_data_validation(validation)
        validation.add(f"{col_letter}{first_data_row}:{col_letter}{last_data_row}")

    return wb
