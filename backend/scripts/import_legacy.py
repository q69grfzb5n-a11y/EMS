#!/usr/bin/env python3
"""Import the real employee roster from the legacy incentives workbook.

Parses the 9 department sheets of `docs/source/Precast Incentives 03-2026.xlsm`
(8 factory departments + the Engs cross-department engineers group, seeded as
its own department — see PLAN.md open question), upserts `employees` keyed by
staff number (Oracle no.), then backfills `full_name_en` by cross-matching
staff numbers against the monthly Oracle attendance export. Idempotent: safe
to re-run as the source file is corrected.

Usage:
    uv run python scripts/import_legacy.py [--file PATH] [--attendance PATH] [--dry-run]
"""

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.common.enums import EmploymentStatus
from app.db.session import SessionLocal
from app.modules.employees.models import Employee
from app.modules.org.models import Department, Position

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSM = REPO_ROOT / "docs" / "source" / "Precast Incentives 03-2026.xlsm"
DEFAULT_ATTENDANCE = REPO_ROOT / "docs" / "source" / "attendance_export_06-2026.xlsx"

# Sheet name -> departments.json code. Confirms the open question in PLAN §11:
# "Engs" is seeded and treated as its own (9th) department, not a cross-dept tag.
DEPT_SHEETS: dict[str, str] = {
    "Management": "MGMT",
    "Production": "PROD",
    "Technical": "TECH",
    "Hollowcore": "HC",
    "QC": "QC",
    "Maintenance": "MAINT",
    "Installation": "INST",
    "SCM": "SCM",
    "Engs": "ENGS",
}

HEADER_ROW = 7
DATA_START_ROW = 8
HEADER_MAX_COL = 12
# Sheets are hand-maintained by HR: column order/blank spacer columns drift between
# sheets and some carry a duplicated "Actual Position Final" header. Column *names*
# from this header row are the only reliable contract — never assume fixed indices.
NAME_COL = "Employee Name"
ORACLE_COL = "Oracle"
CONTRACT_TITLE_COL = "Position as per Oracle"
POSITION_COL = "Actual Position Final"
STATUS_COL = "الحالة"
# The only free-text phrase in the status column that reliably means "no longer
# employed"; every other value (leave notes, "شغال", transfer notes, ...) still
# means an active employee — there is no third status in this system.
FINAL_EXIT_MARKER = "خروج نهائي"


@dataclass
class ParsedRow:
    dept_code: str
    source_row: int
    staff_no: str
    full_name_ar: str
    contract_position_title: str | None
    position_title_raw: str
    employment_status: str


def normalize_staff_no(value: object) -> str:
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    try:
        return str(int(float(text)))
    except ValueError:
        return text


def _header_indices(header_row: list[object]) -> dict[str, list[int]]:
    indices: dict[str, list[int]] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        indices.setdefault(str(h).strip(), []).append(i)
    return indices


def _first_nonempty(vals: list[object], indices: list[int]) -> str | None:
    for i in indices:
        if i < len(vals) and vals[i] not in (None, ""):
            return str(vals[i]).strip()
    return None


def parse_sheet(sheet, dept_code: str) -> tuple[list[ParsedRow], list[dict[str, object]]]:
    """Pure parse: sheet -> (resolvable rows, rows needing a human's attention)."""
    header = [
        c.value
        for c in next(
            sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, max_col=HEADER_MAX_COL)
        )
    ]
    idx = _header_indices(header)
    oracle_idx = idx.get(ORACLE_COL, [])

    rows: list[ParsedRow] = []
    skipped: list[dict[str, object]] = []
    for excel_row, row in enumerate(
        sheet.iter_rows(min_row=DATA_START_ROW, max_col=HEADER_MAX_COL), start=DATA_START_ROW
    ):
        vals = [c.value for c in row]
        oracle = _first_nonempty(vals, oracle_idx)
        if oracle is None:
            continue
        staff_no = normalize_staff_no(oracle)
        name = _first_nonempty(vals, idx.get(NAME_COL, [])) or ""
        contract_title = _first_nonempty(vals, idx.get(CONTRACT_TITLE_COL, []))
        position_raw = _first_nonempty(vals, idx.get(POSITION_COL, []))
        status_text = _first_nonempty(vals, idx.get(STATUS_COL, []))
        employment_status = (
            EmploymentStatus.TERMINATED.value
            if status_text and FINAL_EXIT_MARKER in status_text
            else EmploymentStatus.ACTIVE.value
        )

        if not position_raw:
            skipped.append(
                {
                    "dept": dept_code,
                    "row": excel_row,
                    "staff_no": staff_no,
                    "name": name,
                    "reason": "no 'Actual Position Final' value in source row",
                }
            )
            continue

        rows.append(
            ParsedRow(
                dept_code=dept_code,
                source_row=excel_row,
                staff_no=staff_no,
                full_name_ar=name,
                contract_position_title=contract_title,
                position_title_raw=position_raw,
                employment_status=employment_status,
            )
        )
    return rows, skipped


def import_roster(db: Session, xlsm_path: Path, *, dry_run: bool) -> None:
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    departments_by_code = {d.code: d for d in db.scalars(select(Department)).all()}
    positions_by_title = {p.title_en.strip().lower(): p for p in db.scalars(select(Position)).all()}

    unresolved: list[dict[str, object]] = []
    duplicates: list[dict[str, object]] = []
    seen_staff_no: dict[str, str] = {}
    created = updated = 0

    for sheet_name, dept_code in DEPT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ! sheet '{sheet_name}' not found in workbook, skipping")
            continue
        dept = departments_by_code.get(dept_code)
        if dept is None:
            print(f"  ! department '{dept_code}' not seeded — run seed.py --core first, skip sheet")
            continue

        parsed_rows, skipped = parse_sheet(wb[sheet_name], dept_code)
        unresolved.extend(skipped)

        for row in parsed_rows:
            if row.staff_no in seen_staff_no:
                duplicates.append(
                    {
                        "staff_no": row.staff_no,
                        "name": row.full_name_ar,
                        "first_dept": seen_staff_no[row.staff_no],
                        "duplicate_dept": dept_code,
                        "row": row.source_row,
                    }
                )
                continue

            position = positions_by_title.get(row.position_title_raw.strip().lower())
            if position is None:
                unresolved.append(
                    {
                        "dept": dept_code,
                        "row": row.source_row,
                        "staff_no": row.staff_no,
                        "name": row.full_name_ar,
                        "reason": f"unrecognized position title {row.position_title_raw!r}",
                    }
                )
                continue

            seen_staff_no[row.staff_no] = dept_code
            employee = db.scalars(select(Employee).where(Employee.staff_no == row.staff_no)).first()
            if employee is None:
                db.add(
                    Employee(
                        staff_no=row.staff_no,
                        full_name_ar=row.full_name_ar,
                        department_id=dept.id,
                        position_id=position.id,
                        contract_position_title=row.contract_position_title,
                        employment_status=row.employment_status,
                    )
                )
                created += 1
            else:
                employee.full_name_ar = row.full_name_ar
                employee.department_id = dept.id
                employee.position_id = position.id
                employee.contract_position_title = row.contract_position_title
                employee.employment_status = row.employment_status
                updated += 1

    if dry_run:
        db.rollback()
    else:
        db.commit()

    suffix = " (dry run — rolled back)" if dry_run else ""
    print(f"Roster: {created} created, {updated} updated{suffix}.")
    if unresolved:
        print(f"Unresolved rows needing HR attention ({len(unresolved)}), never skipped silently:")
        for u in unresolved:
            print(f"  ! {u['dept']} row {u['row']}: staff_no={u['staff_no']} name={u['name']!r}")
            print(f"      reason: {u['reason']}")
    if duplicates:
        print(f"Duplicate staff numbers across sheets, kept first occurrence ({len(duplicates)}):")
        for d in duplicates:
            print(
                f"  ! staff_no={d['staff_no']} name={d['name']!r} first seen in "
                f"{d['first_dept']}, duplicate in {d['duplicate_dept']} row {d['row']}"
            )


def enrich_names_from_attendance(db: Session, attendance_path: Path, *, dry_run: bool) -> None:
    wb = openpyxl.load_workbook(attendance_path, data_only=True, read_only=True)
    ws = wb["Employees Time Card Summary Rep"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h).strip(): i for i, h in enumerate(header) if h}
    person_idx = idx.get("Person No.")
    name_idx = idx.get("Name")
    if person_idx is None or name_idx is None:
        print("  ! attendance file missing 'Person No.'/'Name' headers, skipping enrichment")
        return

    name_by_staff_no: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        person_no = vals[person_idx] if person_idx < len(vals) else None
        name = vals[name_idx] if name_idx < len(vals) else None
        if person_no is None or not name:
            continue
        name_by_staff_no[normalize_staff_no(person_no)] = str(name).strip()

    employees = list(db.scalars(select(Employee)))
    enriched = 0
    unmatched = 0
    for employee in employees:
        name_en = name_by_staff_no.get(employee.staff_no)
        if name_en is None:
            unmatched += 1
            continue
        if employee.full_name_en != name_en:
            employee.full_name_en = name_en
            enriched += 1

    if dry_run:
        db.rollback()
    else:
        db.commit()
    print(
        f"Name enrichment: {enriched} employees updated, {unmatched} employees have no "
        f"matching staff number in the attendance export (dry run — rolled back)."
        if dry_run
        else f"Name enrichment: {enriched} employees updated, {unmatched} employees have no "
        f"matching staff number in the attendance export."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--file", type=Path, default=DEFAULT_XLSM, help="Path to the legacy incentives workbook"
    )
    parser.add_argument(
        "--attendance",
        type=Path,
        default=DEFAULT_ATTENDANCE,
        help="Path to the Oracle attendance export, used to backfill full_name_en",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Parse and report only, no DB writes"
    )
    args = parser.parse_args()

    if not args.file.exists():
        parser.error(f"workbook not found: {args.file}")

    db = SessionLocal()
    try:
        print(f"Importing roster from {args.file}{' (dry run)' if args.dry_run else ''}...")
        import_roster(db, args.file, dry_run=args.dry_run)

        if args.attendance.exists():
            print(f"Enriching full_name_en from {args.attendance}...")
            enrich_names_from_attendance(db, args.attendance, dry_run=args.dry_run)
        else:
            print(f"  ! attendance file not found at {args.attendance}, skipping name enrichment")
    finally:
        db.close()


if __name__ == "__main__":
    main()
